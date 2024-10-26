import numpy as np
import os
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle
import time
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO
import logging
import warnings

# 忽略 FutureWarning
warnings.simplefilter(action='ignore', category=FutureWarning)

# 設置日誌記錄
logging.basicConfig(
    filename='monitoring.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# 從環境變數中讀取 LINE Notify 權杖
LINE_NOTIFY_TOKEN = 'QZsIlNXyGMvEbzwrJdXhHdOqxeY8WXzXbAOg3E0m4rN'

def send_line_notify(message, image_bytes=None):
    """
    發送訊息和圖片到 LINE Notify
    :param message: 要發送的文字訊息
    :param image_bytes: 圖片的二進位數據（可選）
    """
    url = "https://notify-api.line.me/api/notify"
    headers = {
        "Authorization": f"Bearer {LINE_NOTIFY_TOKEN}"
    }
    data = {"message": message}
    files = {}

    if image_bytes:
        # 使用 BytesIO 對象傳遞圖片數據
        files = {"imageFile": ("plot.png", image_bytes, "image/png")}

    try:
        response = requests.post(url, headers=headers, data=data, files=files)
        if response.status_code == 200:
            logging.info("成功發送 LINE Notify 訊息。")
        else:
            logging.error(f"發送 LINE Notify 訊息失敗，狀態碼：{response.status_code}")
            logging.error(f"回應內容：{response.text}")
    except Exception as e:
        logging.error(f"發送 LINE Notify 時發生異常：{e}")

def fetch_okx_bybit_data():
    logging.info("開始抓取 OKX 和 Bybit 的現價及資費數據...")
    # API URLs
    okx_ticker_url = "https://www.okx.com/api/v5/market/tickers"
    okx_funding_rate_url = "https://www.okx.com/api/v5/public/funding-rate"
    okx_mark_price_url = "https://www.okx.com/api/v5/public/mark-price"
    bybit_ticker_url = "https://api.bybit.com/v5/market/tickers"

    # API request parameters
    okx_params = {'instType': 'SWAP'}
    bybit_params = {'category': 'linear'}

    # Fetch OKX and Bybit data
    try:
        logging.info("請求 OKX Ticker 資料...")
        okx_response = requests.get(okx_ticker_url, params=okx_params).json()
        logging.info("請求 Bybit Ticker 資料...")
        bybit_response = requests.get(bybit_ticker_url, params=bybit_params).json()
    except requests.RequestException as e:
        logging.error(f"API 請求錯誤: {e}")
        send_line_notify(f"API 請求錯誤: {e}")
        return None, None  # 返回 None 以避免進一步的錯誤

    # Create DataFrames to store data
    okx_df = pd.DataFrame(columns=['Instrument ID', 'OKX Price', 'OKX Mark Price', 'OKX Funding Rate'])
    bybit_df = pd.DataFrame(columns=['Instrument ID', 'Bybit Price', 'Bybit Mark Price', 'Bybit Funding Rate'])

    # Process OKX data
    logging.info("處理 OKX 資料...")
    for item in okx_response.get('data', []):
        inst_id = item.get('instId')
        if 'USDT' in inst_id:
            base_currency = inst_id.replace('-USDT-SWAP', '')
            try:
                ok_price = float(item.get('last', '0'))
            except ValueError:
                ok_price = 0.0

            # Fetch OKX funding rate
            try:
                funding_rate_response = requests.get(okx_funding_rate_url, params={'instId': inst_id}, timeout=10).json()
                ok_funding = float(funding_rate_response['data'][0].get('fundingRate', '0')) if 'data' in funding_rate_response else 0.0
            except (requests.RequestException, ValueError, IndexError):
                ok_funding = 0.0

            # Fetch OKX mark price
            try:
                mark_price_response = requests.get(okx_mark_price_url, params={'instId': inst_id}, timeout=10).json()
                ok_mark = float(mark_price_response['data'][0].get('markPx', '0')) if 'data' in mark_price_response else 0.0
            except (requests.RequestException, ValueError, IndexError):
                ok_mark = 0.0

            # Create a temporary DataFrame for the current row
            temp_okx_df = pd.DataFrame([{
                'Instrument ID': base_currency,
                'OKX Price': ok_price,
                'OKX Mark Price': ok_mark,
                'OKX Funding Rate': ok_funding
            }])

            # 檢查 temp_okx_df 是否不為空且至少有一個有效值
            print(f"Processing OKX data for {inst_id}:")
            print(temp_okx_df)
            if not temp_okx_df.empty and temp_okx_df[['Instrument ID', 'OKX Price', 'OKX Mark Price', 'OKX Funding Rate']].notna().any().any():
                okx_df = pd.concat([okx_df, temp_okx_df], ignore_index=True)
                print(f"Added data for {inst_id}")
            else:
                print(f"跳過空或全為NA的 OKX 資料: {temp_okx_df}")

    # Process Bybit data
    logging.info("處理 Bybit 資料...")
    for item in bybit_response.get('result', {}).get('list', []):
        symbol = item.get('symbol')
        if 'USDT' in symbol:
            base_currency = symbol.replace('USDT', '')
            try:
                bybit_price = float(item.get('lastPrice', '0'))
            except ValueError:
                bybit_price = 0.0
            try:
                bybit_mark = float(item.get('markPrice', '0'))
            except ValueError:
                bybit_mark = 0.0
            try:
                bybit_funding = float(item.get('fundingRate', '0'))
            except ValueError:
                bybit_funding = 0.0

            # Create a temporary DataFrame for the current row
            temp_bybit_df = pd.DataFrame([{
                'Instrument ID': base_currency,
                'Bybit Price': bybit_price,
                'Bybit Mark Price': bybit_mark,
                'Bybit Funding Rate': bybit_funding
            }])

            # 檢查 temp_bybit_df 是否不為空且至少有一個有效值
            print(f"Processing Bybit data for {symbol}:")
            print(temp_bybit_df)
            if not temp_bybit_df.empty and temp_bybit_df[['Instrument ID', 'Bybit Price', 'Bybit Mark Price', 'Bybit Funding Rate']].notna().any().any():
                bybit_df = pd.concat([bybit_df, temp_bybit_df], ignore_index=True)
                print(f"Added data for {symbol}")
            else:
                print(f"跳過空或全為NA的 Bybit 資料: {temp_bybit_df}")

    # Merge DataFrames
    logging.info("合併 OKX 和 Bybit 的資料...")
    merged_df = pd.merge(okx_df, bybit_df, on='Instrument ID', how='inner')
    logging.info(f"合併後的 DataFrame 形狀: {merged_df.shape}")

    if merged_df.empty:
        logging.warning("合併後的資料為空。")
        send_line_notify("⚠️ 抓取的幣種資料為空。")
        return None, None

    # Reorder columns
    ordered_columns = [
        'Instrument ID',
        'Bybit Price',
        'Bybit Mark Price',
        'Bybit Funding Rate',
        'OKX Price',
        'OKX Mark Price',
        'OKX Funding Rate'
    ]
    merged_df = merged_df[ordered_columns]

    # 將 DataFrame 寫入 CSV 文件
    csv_file_path = "USDT_Currencies_Data_Combined.csv"
    logging.info(f"將合併後的資料寫入 CSV 文件: {csv_file_path}")
    merged_df.to_csv(csv_file_path, index=False)

    # 轉換 CSV 文件為 Excel 文件
    excel_file_path = "USDT_Currencies_Data_Combined.xlsx"
    logging.info(f"將 CSV 文件轉換為 Excel 文件: {excel_file_path}")
    df = pd.read_csv(csv_file_path)
    df.to_excel(excel_file_path, index=False)

    # 打開 Excel 文件進行格式化
    logging.info("格式化 Excel 文件...")
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # 設置字體樣式
    font_size = 10
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Arial', size=font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 設置數值格式
    number_format_16 = NamedStyle(name="number_format_16", number_format="0.################")
    for col in 'BCDEFG':
        for cell in ws[col]:
            cell.style = number_format_16

    # 添加公式
    logging.info("添加公式到 Excel 文件...")
    headers = {
        'H1': 'Bybit Fair',
        'I1': 'OKX Fair',
        'J1': 'Bybit_OK',
        'K1': 'Bybit_OK_Abs',
        'L1': 'Bybit_OK_Diff',
        'M1': '利差',
        'N1': '價差',
        'O1': '方向',
        'P1': '總分',
        'Q1': '做多',
        'R1': '做空',
        'S1': '費價比',
        'T1': '3條件都達成',
        'U1': '大於0.05%'
    }
    for cell, value in headers.items():
        ws[cell] = value

    for row in range(2, ws.max_row + 1):
        ws[f'H{row}'] = f'=IF(AND(B{row}<>"N/A",D{row}<>"N/A"), B{row}*(1+D{row}), "N/A")'
        ws[f'I{row}'] = f'=IF(AND(E{row}<>"N/A",G{row}<>"N/A"), E{row}*(1+G{row}), "N/A")'
        ws[f'J{row}'] = f'=IF(AND(H{row}<>"N/A", I{row}<>"N/A", C{row}<>"N/A", F{row}<>"N/A"), (H{row}-I{row})/((C{row}+F{row})/2), "N/A")'
        ws[f'K{row}'] = f'=IF(J{row}<>"N/A", ABS(J{row}), "N/A")'
        ws[f'L{row}'] = f'=IF(AND(D{row}<>"N/A", G{row}<>"N/A"), ABS(D{row}-G{row}), "N/A")'
        ws[f'M{row}'] = f'=IF(L{row}<>"N/A", IF(L{row}>0.0003, 1, 0), "N/A")'
        ws[f'N{row}'] = f'=IF(AND(K{row}<>"N/A", L{row}<>"N/A"), IF(K{row}-L{row}>0.0002, 1, 0), "N/A")'
        ws[f'O{row}'] = f'=IF(AND(J{row}<>"N/A", D{row}<>"N/A", G{row}<>"N/A"), IF(J{row}*(D{row}-G{row})>0, 1, 0), "N/A")'
        ws[f'P{row}'] = f'=IF(AND(M{row}<>"N/A", N{row}<>"N/A", O{row}<>"N/A"), M{row} + N{row} + O{row}, "N/A")'
        ws[f'Q{row}'] = f'=IF(AND(D{row}<>"N/A", G{row}<>"N/A"), IF(D{row}<G{row}, "Bybit", "OKEX"), "N/A")'
        ws[f'R{row}'] = f'=IF(AND(D{row}<>"N/A", G{row}<>"N/A"), IF(D{row}>G{row}, "Bybit", "OKEX"), "N/A")'
        ws[f'S{row}'] = f'=IF(AND(G{row}>D{row}, F{row}>C{row}), 1, IF(AND(G{row}<D{row}, F{row}<C{row}), 1, 0))'
        ws[f'T{row}'] = f'=IF(AND(P{row}<>"N/A", S{row}<>"N/A", U{row}<>"N/A"), P{row} + S{row} + U{row}, "N/A")'
        ws[f'U{row}'] = f'=IF(L{row}>0.0005, 1, 0)'

    # 設置列寬
    column_widths = {
        'A': 13, 'B': 11, 'C': 16, 'D': 18, 'E': 10, 'F': 16, 'G': 22, 'H': 13, 'I': 13, 'J': 14, 'K': 13, 'L': 13,
        'M': 6,
        'N': 6, 'O': 6, 'P': 6, 'Q': 8, 'R': 8, 'S': 7, 'T': 13, 'U': 10
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 保存 Excel 文件
    logging.info(f"保存格式化後的 Excel 文件: {excel_file_path}")
    wb.save(excel_file_path)

    # 刪除 CSV 文件
    logging.info(f"刪除暫存的 CSV 文件: {csv_file_path}")
    os.remove(csv_file_path)

    # 讀取 Excel 文件
    logging.info("讀取格式化後的 Excel 文件以進行進一步處理...")
    df = pd.read_excel(excel_file_path)

    # 計算條件
    logging.info("計算過濾條件...")
    df['條件1'] = np.abs(df['Bybit Funding Rate'] - df['OKX Funding Rate']) > 0.0003
    df['條件2'] = ((np.abs((df['Bybit Price'] * (1 + df['Bybit Funding Rate'])) -
                           (df['OKX Price'] * (1 + df['OKX Funding Rate']))) /
                        np.abs((df['Bybit Mark Price'] + df['OKX Mark Price']) / 2)) -
                   np.abs(df['Bybit Funding Rate'] - df['OKX Funding Rate'])) > 0.0002
    df['條件3'] = ((df['Bybit Price'] * (1 + df['Bybit Funding Rate'])) -
                   (df['OKX Price'] * (1 + df['OKX Funding Rate']))) * (
                              df['Bybit Funding Rate'] - df['OKX Funding Rate']) > 0
    df['條件4'] = (((df['OKX Mark Price'] < df['Bybit Mark Price']) &
                    (df['OKX Funding Rate'] < df['Bybit Funding Rate'])) |
                   ((df['OKX Mark Price'] > df['Bybit Mark Price']) &
                    (df['OKX Funding Rate'] > df['Bybit Funding Rate'])))
    df['條件5'] = np.abs(df['Bybit Funding Rate'] - df['OKX Funding Rate']) > 0.0005

    # 過濾符合條件的行
    logging.info("過濾符合條件的幣種...")
    filtered_df = df[(df['條件1']) & (df['條件2']) & (df['條件3']) & (df['條件4']) & (df['條件5'])].copy()

    return filtered_df, excel_file_path

def fetch_and_process_funding_data(currencies):
    logging.info("開始抓取並處理歷史資費數據...")
    currency_data = {}

    for currency in currencies:
        try:
            currency = currency.strip()
            symbol = f"{currency}USDT"
            instId = f"{currency}-USDT-SWAP"

            end_time = int(time.time() * 1000)
            start_time = int((datetime.now() - timedelta(days=90)).timestamp() * 1000)
            limit = 200

            dates = [(datetime.now() - timedelta(days=day)).strftime('%Y-%m-%d') for day in range(90)]

            all_okx_data = []
            for date in dates:
                okx_data = fetch_okx_funding_rate_data(instId, date)
                if okx_data:
                    all_okx_data.extend(process_okx_data(okx_data))
                time.sleep(0.2)  # 防止過快請求

            all_bybit_data = []
            while True:
                bybit_data = fetch_bybit_funding_rate_data(symbol, start_time, end_time, limit)
                if not bybit_data:
                    break
                all_bybit_data.extend(process_bybit_data(bybit_data))
                end_time = int(bybit_data[-1]['fundingRateTimestamp']) - 1
                time.sleep(0.2)  # 防止過快請求

            currency_data[currency] = (all_bybit_data, all_okx_data)
            logging.info(f"已抓取 {currency} 的資費數據。")
        except Exception as e:
            logging.error(f"處理 {currency} 時出錯: {e}")
            send_line_notify(f"處理 {currency} 時出錯: {e}")

    return currency_data

def fetch_bybit_funding_rate_data(symbol, start_time, end_time, limit):
    url = "https://api.bybit.com/v5/market/funding/history"
    params = {
        'symbol': symbol,
        'startTime': start_time,
        'endTime': end_time,
        'limit': limit,
        'category': 'linear'
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()  # 檢查 HTTP 狀態碼
        data = response.json()
        if data.get('retCode') == 0 and 'result' in data:
            return data['result']['list']
        else:
            logging.error(f"Bybit 抓取數據錯誤: {data.get('retMsg', 'Unknown error')}")
            return []
    except requests.RequestException as e:
        logging.error(f"Bybit API 請求錯誤: {e}")
        send_line_notify(f"Bybit API 請求錯誤: {e}")
        return []

def fetch_okx_funding_rate_data(instId, date):
    url = 'https://www.okx.com/api/v5/public/funding-rate-history'
    params = {
        'instId': instId,
        'limit': 1000,
        'before': int(time.mktime(datetime.strptime(date, '%Y-%m-%d').timetuple()) * 1000)
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 429:
            logging.warning(f"OKX 速率限制，等待 60 秒後重試...")
            time.sleep(60)
            return fetch_okx_funding_rate_data(instId, date)
        response.raise_for_status()  # 檢查 HTTP 狀態碼
        response_data = response.json()
        if 'data' in response_data:
            return response_data['data']
        else:
            logging.error(f"OKX 抓取數據錯誤 (日期 {date}): {response_data.get('error_message', 'Unknown error')}")
            return []
    except requests.RequestException as e:
        logging.error(f"OKX API 請求錯誤: {e}")
        send_line_notify(f"OKX API 請求錯誤: {e}")
        return []

def process_bybit_data(data):
    return [{
        'Timestamp': time.strftime('%Y-%m-%d %H:%M:%S',
                                   time.localtime(int(entry['fundingRateTimestamp']) / 1000 + 8 * 3600)),
        'Bybit Funding Rate': float(entry['fundingRate']) * 100
    } for entry in data]

def process_okx_data(data):
    return [{
        'Timestamp': datetime.fromtimestamp(int(entry['fundingTime']) / 1000).strftime('%Y-%m-%d %H:%M:%S'),
        'OKX Realized Rate': float(entry['realizedRate']) * 100
    } for entry in data]

def save_to_excel_funding_data(currency_data, file_path):
    logging.info(f"保存歷史資費數據到 Excel 文件: {file_path}")
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        for currency, (bybit_data, okx_data) in currency_data.items():
            logging.info(f"處理 {currency} 的歷史資費數據...")
            bybit_df = pd.DataFrame(bybit_data)
            okx_df = pd.DataFrame(okx_data)

            bybit_df['Timestamp'] = pd.to_datetime(bybit_df['Timestamp'])
            okx_df['Timestamp'] = pd.to_datetime(okx_df['Timestamp'])

            combined_df = pd.merge(okx_df, bybit_df, on='Timestamp', how='outer')
            combined_df['Bybit Funding Rate'] = combined_df['Bybit Funding Rate'].fillna(0)
            combined_df['OKX Realized Rate'] = combined_df['OKX Realized Rate'].fillna(0)
            combined_df['Calculated Result'] = (combined_df['OKX Realized Rate'] * -1) + combined_df[
                'Bybit Funding Rate']
            combined_df.drop_duplicates(inplace=True)

            # Save data to separate sheet
            combined_df.to_excel(writer, sheet_name=currency, index=False)
            logging.info(f"已保存 {currency} 的數據至 Excel 工作表。")

    logging.info(f"歷史資費數據已保存至 {file_path}")

def plot_and_send(df, currency):
    logging.info(f"生成並準備發送 {currency} 的計算結果圖表...")
    plt.figure(figsize=(16, 8))  # 調整圖表大小

    # 確保 'Timestamp' 是 datetime 類型
    if not np.issubdtype(df['Timestamp'].dtype, np.datetime64):
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])

    # Determine color for bars and edges
    colors = df['Calculated Result'].apply(lambda x: 'red' if x > 0 else 'green')
    edgecolors = df['Calculated Result'].apply(lambda x: 'red' if x > 0 else 'green')

    # Calculate number of positive and negative bars
    red_count = (df['Calculated Result'] > 0).sum()
    green_count = (df['Calculated Result'] <= 0).sum()

    # Plot bar chart
    plt.bar(df['Timestamp'], df['Calculated Result'], color=colors, edgecolor=edgecolors, width=0.2)

    # Title and labels
    plt.title(f'Calculated Result for {currency} Over Time', fontsize=18)
    plt.xlabel('Time', fontsize=14)
    plt.ylabel('Calculated Result', fontsize=14)

    # Set x-axis major locator and formatter
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=2))
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.gcf().autofmt_xdate()

    # Adjust y-axis range
    min_val = df['Calculated Result'].min()
    max_val = df['Calculated Result'].max()
    buffer = (max_val - min_val) * 0.1 if (max_val - min_val) != 0 else 1
    plt.ylim(min_val - buffer, max_val + buffer)

    # Add grid
    plt.grid(True, linestyle='--', alpha=0.7)

    # Add annotation for red and green bar counts
    plt.text(df['Timestamp'].iloc[-1], max_val, f'Red Bars: {red_count}\nGreen Bars: {green_count}',
             fontsize=12, color='black', ha='right', va='top', bbox=dict(facecolor='white', alpha=0.5))

    # Save plot to BytesIO buffer
    buf = BytesIO()
    plt.savefig(buf, format='png')
    plt.close()  # Close the plot without displaying it
    buf.seek(0)
    image_bytes = buf.read()

    return image_bytes

def main():
    logging.info("開始實時監控程序。")
    while True:
        try:
            # 抓取和處理 OKX 和 Bybit 的價格及資費數據
            filtered_df, combined_excel_path = fetch_okx_bybit_data()

            if filtered_df is None or combined_excel_path is None:
                logging.warning("抓取數據時出現錯誤，跳過此次循環。")
                send_line_notify("⚠️ 抓取數據時出現錯誤，請檢查程式碼。")
            elif filtered_df.empty:
                logging.info("沒有任何幣種符合條件。")
                send_line_notify("📉 沒有任何幣種符合條件。")
            else:
                # 提取符合條件的幣種列表
                filtered_currencies = filtered_df['Instrument ID'].unique().tolist()
                logging.info(f"已抓取符合條件的幣種: {filtered_currencies}")

                # 抓取並處理歷史資費數據
                currency_data = fetch_and_process_funding_data(filtered_currencies)

                # 保存歷史資費數據到 Excel 文件
                file_path = 'funding_data.xlsx'
                save_to_excel_funding_data(currency_data, file_path)
                logging.info(f"資費數據已保存至 {file_path}")

                # 使用 .loc 避免 SettingWithCopyWarning
                filtered_df.loc[:, '結果1'] = np.where(filtered_df['Bybit Funding Rate'] < filtered_df['OKX Funding Rate'], 'Bybit',
                                                       'OKEX')
                filtered_df.loc[:, '結果2'] = np.where(filtered_df['Bybit Funding Rate'] > filtered_df['OKX Funding Rate'], 'Bybit',
                                                       'OKEX')

                # 添加調試訊息以確認 '結果1' 和 '結果2' 是否存在
                logging.debug("確認 '結果1' 和 '結果2' 是否已添加到 filtered_df...")
                logging.debug(f"Columns: {filtered_df.columns}")
                logging.debug(f"Head:\n{filtered_df.head()}")

                # 準備並發送每個幣種的訊息和圖表
                for index, row in filtered_df.iterrows():
                    # 獲取當前日期和時間（精確到分鐘）
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M')

                    # 構建每個幣種的訊息
                    message = (
                        f"\n"
                        f"🕒  {current_time}\n\n"
                        f"🪙 幣種: {row['Instrument ID']}\n\n"
                    )

                    # 根據 '結果1' 和 '結果2' 獲取對應的數據
                    if row['結果1'] == 'OKEX':
                        mark_price_1 = f"{row['OKX Mark Price']:.4f}"
                        funding_rate_1 = f"{row['OKX Funding Rate'] * 100:.4f}%"
                    else:
                        mark_price_1 = f"{row['Bybit Mark Price']:.4f}"
                        funding_rate_1 = f"{row['Bybit Funding Rate'] * 100:.4f}%"

                    if row['結果2'] == 'OKEX':
                        mark_price_2 = f"{row['OKX Mark Price']:.4f}"
                        funding_rate_2 = f"{row['OKX Funding Rate'] * 100:.4f}%"
                    else:
                        mark_price_2 = f"{row['Bybit Mark Price']:.4f}"
                        funding_rate_2 = f"{row['Bybit Funding Rate'] * 100:.4f}%"

                    # 添加結果到訊息
                    message += (
                        f"🔼 做(多)交易所: {row['結果1']}\n"
                        f"💰 價格: {mark_price_1}\n"
                        f"📊 資費: {funding_rate_1}\n"
                        f"🔽 做(空)交易所: {row['結果2']}\n"
                        f"💰 價格: {mark_price_2}\n"
                        f"📊 資費: {funding_rate_2}\n"
                    )

                    # 發送幣種資訊到 LINE 並附上圖表圖片
                    currency = row['Instrument ID']
                    try:
                        funding_df = pd.read_excel(file_path, sheet_name=currency)
                        # 計算 'Calculated Result' 以供繪圖
                        if 'Calculated Result' not in funding_df.columns:
                            funding_df['Calculated Result'] = (funding_df['OKX Realized Rate'] * -1) + funding_df[
                                'Bybit Funding Rate']

                        # 生成圖表
                        image_bytes = plot_and_send(funding_df, currency)

                        # 發送訊息和圖片到 LINE
                        send_line_notify(message, image_bytes=image_bytes)
                        logging.info(f"已發送 {currency} 的 LINE 通知。")
                    except Exception as e:
                        error_message = f"⚠️ 無法生成 {currency} 的圖表或發送訊息：{e}"
                        logging.error(error_message)
                        send_line_notify(error_message)

            # 等待下一次循環
            logging.info("等待 5 分鐘後進行下一次監控。")
            time.sleep(300)  # 300 秒 = 5 分鐘

        except KeyboardInterrupt:
            logging.info("收到中斷信號，終止監控程序。")
            print("監控程序已終止。")
            break
        except Exception as e:
            logging.error(f"在主循環中發生未處理的異常: {e}")
            send_line_notify(f"⚠️ 在主循環中發生未處理的異常: {e}")
            time.sleep(60)  # 等待一分鐘後重試

    logging.info("實時監控程序已結束。")

if __name__ == "__main__":
    main()
